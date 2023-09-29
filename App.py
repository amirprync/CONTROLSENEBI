import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill

def highlight_rows(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        wb = writer.book
        ws = wb.active
        red_fill = PatternFill(start_color="FFFF00",
                               end_color="FFFF00", fill_type="solid")
        
        for row in range(2, df.shape[0] + 2):
            comitente = ws.cell(row=row, column=1).value
            codigo_caja = ws.cell(row=row, column=2).value
            cantidad_a_transferir = ws.cell(row=row, column=5).value
            
            condition = ((df['Comitente - Número'] == comitente) & 
                         (df['Instrumento - Código caja'] == codigo_caja) & 
                         (df['Transferencia - Cantidad a Transferir'] == cantidad_a_transferir))
            
            if df[condition].shape[0] > 1:
                for col in range(1, df.shape[1] + 1):
                    ws.cell(row=row, column=col).fill = red_fill
    output.seek(0)
    return output

# Interfaz Streamlit
st.title("Resaltador de Filas en Excel")

uploaded_file = st.file_uploader("Elige un archivo Excel", type=['xlsx'])

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    st.write("Datos cargados:")
    st.write(df)
    
    if st.button('Resaltar Filas'):
        output = highlight_rows(df)
        st.download_button(
            label="Descargar archivo Excel resaltado",
            data=output,
            file_name='datos_resaltados.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
